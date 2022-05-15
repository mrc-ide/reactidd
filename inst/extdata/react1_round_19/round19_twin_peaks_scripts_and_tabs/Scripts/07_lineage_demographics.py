import pandas as pd
import numpy as np
import openpyxl
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats.proportion import proportion_confint

lineageDf = pd.read_csv("E:/dt20/data/rep18_lineage.csv")
lineageDf["age_group_char3"].replace({"5-11":"05-11"},inplace=True)
lineageDf["urban"].replace({0:"rural", 1:"urban"},inplace=True)
lineageDf = lineageDf[lineageDf["react_lineage"].str.contains("BA.1|BA.1.1|BA.2")]

# continuous columns:
# age, nchild, imd, hh_size

# categorical columns:
# sex, age_group_char3, region, urban, employment, ethnic

def propCategorical(df):
    
    return (
        df        
        .apply(lambda x: round(100*x / float(x.sum()),1))
        .rename(columns={"BA.1":"%BA.1",
                         "BA.1.1":"%BA.1.1",
                         "BA.2":"%BA.2"})
    )

def propConfint(df):
    # needs to be counts Df    
    variantsList = ["BA.1", "BA.1.1", "BA.2"]
    indexNameList = ["BA.1pLower95ci", "BA.1.1pLower95ci", "BA.2pLower95ci",
                     "BA.1pUpper95ci", "BA.1.1pUpper95ci", "BA.2pUpper95ci"]

    holderList = []
    for cat in df.index:
        lowDf, upDf = proportion_confint(
            count = df.loc[cat], nobs = df.sum(axis=0))
        tempDf = pd.concat([lowDf, upDf])
        print(tempDf)
        tempDf.index = indexNameList
        holderList.append(tempDf)

    fullCi = pd.concat(holderList, axis=1).transpose()
    df = pd.concat([df,fullCi.set_index(df.index)], axis=1)
    
    df = (
        df
        .drop(variantsList, axis=1)
        .drop_duplicates()
    )
    
    return df    
    

def categoricalCounts(df, categoryCol):
    print(categoryCol)
    
    outDf = (
        df[[categoryCol, "react_lineage"]]
        .pivot_table(index=categoryCol,
                     columns="react_lineage",
                     aggfunc="size",
                     fill_value=0)
        .reset_index()
        .rename_axis(None, axis=1)
        .rename(columns={categoryCol:"category"})
        .sort_values(by=["category"])
        .set_index("category")
        )

    outDf = pd.concat(
        [outDf,
         propCategorical(outDf),
         propConfint(outDf)
         ], axis=1)

    return outDf


def continuousSummary(df, contColList):
    contColList.append("react_lineage")    
    
    outDf = (
        df[contColList]
        .pivot_table(columns="react_lineage", aggfunc=[np.mean, min, max])
        .reset_index()
        .rename(columns={"index":"category"})
        .rename_axis([None, None], axis=1)        
        )

    return outDf


def chi2Test(df):
    # needs to be counts df
    return round(stats.chi2_contingency(df)[1],3)

def testDifference(df, colList, testType):

    pVals =[]

    if testType == "continuous":
        for col in colList:
            pVals.append(
                stats.f_oneway(
                    *lineageDf[[col, "react_lineage"]]
                    .groupby(["react_lineage"])
                    .apply(lambda x: list(x[col])))[1]
                )
    
    return pVals

def prettifyCols(df, prettyType, originalDf=None):
    variantsList = ["BA.1", "BA.1.1", "BA.2"]

    if prettyType == "categorical":
        for variant in variantsList:
            df[variant + " (% [95%CI])"] = \
                       df[variant].astype(str) + \
                       " (" + df["%"+variant].astype(str) + \
                       " [" + round(df[variant+"pLower95ci"]*100,1).astype(str) + \
                       " - " + round(df[variant+"pUpper95ci"]*100,1).astype(str) + \
                       "])"

        desiredCols = [x + " (% [95%CI])" for x in variantsList]       

        

    if prettyType == "continuous":
        for variant in variantsList:
            df[variant] = \
                round(df["mean"][variant], 1).astype(str) + \
                " (" + df["min"][variant].astype(str) + \
                " - " + df["max"][variant].astype(str) + ")"

        desiredCols = [x for x in variantsList]
        desiredCols.insert(0, "category")
        df["category"] = df["category"].replace({                                
                                "hh_size" : "household size",
                                "imd_quintile" : "deprivation index quintile",
                                "nchild"  : "number of children"})
        df["category"] = ["Mean " + x + " (range)" for x in df["category"]]
        df = df.droplevel(1, axis=1)
        
        try:            
            df["p-values"] = testDifference(originalDf, testType="continuous",
                                            colList = ["age", "hh_size", "imd_quintile", "nchild"])
            desiredCols.append("p-values")
        except:
            pass
    
    return df[desiredCols]


#### main ####

categoricalColumnsList = ["gender_char", "age_group_char3",
                          "region", "ethnic_new_char",
                          "urban", "work_new_char"]

continuousColumnsList = ["age", "hh_size", "imd_quintile", "nchild"]

appended_counts = []
chi_pvals = []

for cat in categoricalColumnsList:
    appended_counts.append(categoricalCounts(lineageDf, cat))
    chi_pvals.append(chi2Test(categoricalCounts(lineageDf, cat)))

# appended_counts = pd.concat(appended_counts)


writer = pd.ExcelWriter("../Tables/variant_demographics.xlsx", engine="xlsxwriter")

rowList = [2, 5, 15, 25, 31, 34]

for idx, df in enumerate(appended_counts):
    (prettifyCols(df, "categorical")).to_excel(writer, index=False, header=False,
                                               sheet_name="1", startrow=rowList[idx],
                                               startcol=3)

(prettifyCols(continuousSummary(lineageDf, continuousColumnsList), "continuous", lineageDf)).to_excel(writer, index=False, sheet_name="2")

writer.save()

## consider use openpyxl for whole writing, to prevent reformatting

book  = openpyxl.load_workbook("../Tables/variant_demographics.xlsx")
sheet = book.get_sheet_by_name("1")
sheet.cell(column = 7, row = 3).value = chi_pvals[0]
sheet.cell(column = 7, row = 6).value = chi_pvals[1]
sheet.cell(column = 7, row = 16).value = chi_pvals[2]
sheet.cell(column = 7, row = 26).value = chi_pvals[3]
sheet.cell(column = 7, row = 32).value = chi_pvals[4]
sheet.cell(column = 7, row = 35).value = chi_pvals[5]
book.save("../Tables/variant_demographics.xlsx")

pd.concat(appended_counts).to_csv("E:/custom_marc/data/interim_tables/R18variantCI.csv")

